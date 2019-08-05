# -*- coding: utf-8 -*-
"""
Created on Tue Jul 30 09:31:39 2019

@author: MAX FUNG E202770
"""

'''
DEVELOPMENT PURPOSES ONLY

cd /d C:\engapps\Anaconda2
python C:\Users\E202770\Documents\PXL\PXL.py

cd /d C:\engapps\Anaconda2\scripts

pyinstaller --onedir --noconsole --icon=C:\Users\E202770\Documents\PXL\icon.ico C:\Users\E202770\Documents\PXL\PXL.py

pyi-makespec --onedir --noconsole --icon=C:\Users\E202770\Documents\PXL\icon.ico C:\Users\E202770\Documents\PXL\PXL.py
pyinstaller PXL.spec

'''


# Necessary imports

import sys, os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string, range_boundaries
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font

# PyQt4 Elements

from PyQt5.QtWidgets import QInputDialog, QGraphicsProxyWidget, QGraphicsView, QGraphicsScene, QGraphicsTextItem, QGraphicsOpacityEffect, QFileDialog, QComboBox, QPushButton, QHBoxLayout, QFrame, QVBoxLayout, QApplication, QWidget, QLabel, QDesktopWidget, QLineEdit
from PyQt5.QtGui import QIcon, QPixmap, QFont, QMovie, QFontDatabase
from PyQt5.QtCore import Qt, QCoreApplication, QObject, QThread, pyqtSignal, pyqtSlot, QPropertyAnimation, QTimer


# Relative Path Generator for PyInstaller
# Solution by Jonathon Reinhart on Stack Overflow
# https://stackoverflow.com/questions/7674790/bundling-data-files-with-pyinstaller-onefile

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

logo = resource_path("img\\logo.png")
icon = resource_path("img\\icon.png")
loader = resource_path("img\\loader.gif") # https://loading.io/animation/icon/
blip = resource_path("img\\blip.png")
txt = resource_path("settings.txt") 
fnt = resource_path("fonts/bahnschrift.ttf")
                
                
# By Raures on Python-Forum.io
def readTxt():  
    if os.path.exists(txt):
        with open(txt, "r") as f:
            savePath = f.read()
            print savePath
            f.close()
    else:
        with open(txt, "w") as f:
            savePath = os.path.expanduser("~") + "\\Documents"
            f.write(savePath)
            f.close()
            print 'No Path Exists'
    return savePath

    
# File drop section functionality

class Box(QGraphicsView):
    
    working = pyqtSignal(int)
    
    head = "<font face='Bahnschrift' size=8 color=gray align='center'>"
    tail = "</font>"
    mainMsg = head+"Drag and drop the <b>POs Actions Constraints Spreadsheet</b> (<b>*.xlsx</b>) file here"+tail
    font = QFont()
    font.setFamily("Bahnschrift")
    font.setPixelSize(6)
    fPN = None
    
    def __init__(self):
        super(Box, self).__init__()
        
        self.dragDrop = QGraphicsScene()
        self.setScene(self.dragDrop)
        self.setStyleSheet('Background:transparent;')
        self.setText(self.mainMsg)
        

    def setText(self, string):
        selItems = self.dragDrop.items()
        if selItems:
            self.dragDrop.clear()
            self.dragDrop = QGraphicsScene()
            self.setScene(self.dragDrop)
        self.msg = QGraphicsTextItem()
        self.msg.setHtml(string)
        self.msg.setFont(self.font)
        self.dragDrop.addItem(self.msg)
 
    def dragEnterEvent(self,e):
        for url in e.mimeData().urls():
            self.path = url.toLocalFile()
        if '.xlsx' in self.path:
            cont = self.msg.toPlainText()
            if "file here" not in cont:
                self.setText(self.mainMsg)
                self.resetTimer.stop()
            e.accept()
        else:
            self.unfade()
            self.setText(self.head+"Incompatible file type, must be <b>.xlsx</b> only."+self.tail)         
            self.resetTimer = QTimer()
            self.resetTimer.setInterval(3000)
            self.resetTimer.setSingleShot(True)
            self.resetTimer.timeout.connect(self.reset)
            self.resetTimer.start()

    def dragMoveEvent(self, e):
        for url in e.mimeData().urls():
            self.path = url.toLocalFile()
        if '.xlsx' in self.path:
            e.accept()
        else:
            e.ignore()
 
    def dropEvent(self, e):
        QCoreApplication.processEvents()
        self.setAcceptDrops(False)
        QCoreApplication.processEvents()
        self.setText(self.head+"Preparing the file for analysis..."+self.tail)
        QCoreApplication.processEvents()
        for url in e.mimeData().urls():
            self.url = url.toLocalFile()
            QCoreApplication.processEvents()
            if 'PO' in self.url and 'PXL_' not in self.url:
                print 'passes'
                try:
                    QTimer.singleShot(1,self.spin)
                    QCoreApplication.processEvents()
                    self.threadObject = ThreadPXL(self)
                    self.threadObject.percent.connect(self.updateProgress)
                    self.threadObject.createWorkerThread()
                    self.data_ready()
                except:
                    self.openError()
            else:
                print 'fails'
                self.formatError()

                
#                
    def loadBlip(self):
        self.blip = QPixmap(blip)
        self.setPixmap(self.blip)
        self.show()

    def spin(self):
        self.unfade()
        self.dragDrop.clear()
        self.dragDrop = QGraphicsScene()
        self.setScene(self.dragDrop)
        self.loader = QMovie(loader)
        self.gifSpin = QLabel()
        self.gifSpin.setMovie(self.loader)
        self.loader.start()
        newProxy = QGraphicsProxyWidget()
        newProxy.setWidget(self.gifSpin)
        self.dragDrop.addItem(newProxy)
        self.msg = QGraphicsTextItem()
        self.msg.setHtml(self.head+'0%'+self.tail)
        self.msg.setFont(self.font)
        self.msg.setPos(18,70)
        self.dragDrop.addItem(self.msg)
        
    def updateProgress(self,num):
        if num == 'Opening':
            self.dragDrop.removeItem(self.msg)
            self.msg = QGraphicsTextItem()
            self.msg.setFont(self.font)
            self.msg.setHtml(self.head+num+self.tail)
            self.dragDrop.addItem(self.msg)
            self.msg.setPos(2,70)
            self.working.emit(0)
        else:
            self.msg.setHtml(self.head+num+self.tail)
            self.working.emit(1)
        
        

    def data_ready(self):
        print 'Finished\n\n'
        self.unfade()
        self.setText(self.head+"PXL Analysis is complete. To run another analysis, drag and drop a new file here."+self.tail)
        self.setAcceptDrops(True)
        self.resetTimer = QTimer()
        self.resetTimer.setInterval(20000)
        self.resetTimer.setSingleShot(True)
        self.resetTimer.timeout.connect(self.reset)
        self.resetTimer.start()
        
    def openError(self):
        print 'An unexpected error occurred.\n\n'
        self.unfade()
        self.setText(self.head+"<b>ERROR:</b>   There is currently a PXL spreadsheet open under the same name. Please close it and try again."+self.tail)
        self.setAcceptDrops(True)
        self.resetTimer = QTimer()
        self.resetTimer.setInterval(20000)
        self.resetTimer.setSingleShot(True)
        self.resetTimer.timeout.connect(self.reset)
        self.resetTimer.start()
        self.working.emit(0)
        
    def formatError(self):
        print 'An unexpected error occurred.\n\n'
        self.unfade()
        self.setText(self.head+"<b>ERROR:</b>   The file is not supported. Please check that it is a <b>POs Actions Constraints Spreadsheet</b> and try again."+self.tail)
        self.setAcceptDrops(True)
        self.resetTimer = QTimer()
        self.resetTimer.setInterval(20000)
        self.resetTimer.setSingleShot(True)
        self.resetTimer.timeout.connect(self.reset)
        self.resetTimer.start()
        self.working.emit(0)
        
        
    def reset(self):
        self.unfade()
        self.setText(self.mainMsg)
        self.setAcceptDrops(True)
        
    def fade(self):
        self.effect = QGraphicsOpacityEffect()
        self.setGraphicsEffect(self.effect)
    
        self.animation = QPropertyAnimation(self.effect, b"opacity")
        self.animation.setDuration(1000)
        self.animation.setStartValue(1)
        self.animation.setEndValue(0)
        self.animation.start()

    def unfade(self):
        self.effect = QGraphicsOpacityEffect()
        self.setGraphicsEffect(self.effect)
    
        self.animation = QPropertyAnimation(self.effect, b"opacity")
        self.animation.setDuration(1000)
        self.animation.setStartValue(0)
        self.animation.setEndValue(1)
        self.animation.start()
        
        
        
# PXL Main Object

# Object with threading functionality

class ThreadPXL(QObject):
    
    percent = pyqtSignal(str)

    def __init__(self, Box, parent=None):
        super(self.__class__, self).__init__(parent)
        self.url = Box.url
        self.fPN = Box.fPN

    def createWorkerThread(self):

        # Setup the worker object and the worker_thread.
        
#        def slot(arg='finished'): print(arg)    #Slows process, might delete
        self.worker = PXLCore(self.url, self.fPN)
        self.worker.percent.connect(self.passPercent)
#        self.worker.progress.connect(slot)
#        self.worker.finished.connect(slot)
        self.worker.run()
        
    def passPercent(self,perc):
        self.percent.emit(perc)
#        print perc
        
        # Connect any worker signals
#
#    def forceWorkerReset(self):      
#        if self.worker_thread.isRunning():
#            print('Terminating thread.')
#            self.worker_thread.terminate()
#
#            print('Waiting for thread termination.')
#            self.worker_thread.wait()
#
#            self.signalStatus.emit('Idle.')
#
#            print('building new working object.')
#            self.createWorkerThread()
#
#    def forceWorkerQuit(self):
#        if self.worker_thread.isRunning():
#            self.worker_thread.terminate()
#            self.worker_thread.wait()


# PXL Core Functionality WorkerObject

class PXLCore(QThread):
    
    percent = pyqtSignal(str)
    
    def __init__(self, url, fPN, parent=None):
        super(self.__class__,self).__init__(parent)
        self.url = url
        self.fPN = fPN

    @pyqtSlot()
    def run(self):
        
        # List all files and directories in current directory
        
        xlsx = self.url
        savePath = readTxt()
        
#         Save path for PXL analysis: savePath

        newTitle = "PXL_" + xlsx.split("/",-1)[-1]
        newUrl = savePath + '/' + newTitle

        # Load in the workbook
        
        wb = load_workbook(filename=xlsx, data_only=True) # Constraint: xlsx name
        sheet = wb.active
        
        # Create PXL Analysis workbook
        
        pwb = Workbook()
        pwb.remove(pwb.active)
        psheet = []
        for catg in ['Overview','Conflicts','Matches','OSPs']:
            pwb.create_sheet(title=catg)
            psheet.append(pwb[catg])
        
        # Add titles to sheets
        
        min_col, min_row, max_col, max_row = range_boundaries("A:AK")
        
        for cell in sheet[1][min_col-1:max_col]:
            if 'Program' in cell.value:
                c1 = cell.column
            elif 'Project Number' == cell.value:
                c2 = cell.column
            elif 'Task Number' == cell.value:
                c3 = cell.column
            elif 'PO Number' == cell.value:
                c4 = cell.column
            elif 'Item Description' in cell.value:
                c5 = cell.column
            elif 'Item' == cell.value:
                c6 = cell.column
            elif 'Item Revision' in cell.value:
                c7 = cell.column
            elif 'PPAR Dwg Rev' == cell.value:
                c8 = cell.column
            elif 'PPAR ID' == cell.value:
                c9 = cell.column
            elif 'PPAR Rev' == cell.value:
                c10 = cell.column
            elif 'PPAR Seq' == cell.value:
                c11 = cell.column
            elif 'PPAR' in cell.value:
                c12 = cell.column
            elif 'Buyer Name' in cell.value:
                c13 = cell.column
            elif 'Supplier Name' in cell.value:
                c14 = cell.column
            elif 'QE' in cell.value:
                c15 = cell.column
            
        c = [c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12, c13, c14, c15]
        
        # Define total row length of sheet
        
        len_1 = len(sheet[c7])
        len_2 = len(sheet[c12])
        len_max_range = max(len_1, len_2) + 1
        
        for i in range(1,4):
            for g in range(0,len(c)):
                psheet[i].cell(row=1,column=g+1).value = sheet.cell(row=1,column=column_index_from_string(c[g])).value
                psheet[i].cell(row=1, column = g+1).font = Font(name = 'Bahnschrift', size="14")
                psheet[i].cell(row=1, column = g+1).alignment = Alignment(horizontal='left', vertical='center')
            psheet[i].column_dimensions['A'].width = 40
            psheet[i].column_dimensions['B'].width = 20
            psheet[i].column_dimensions['C'].width = 20
            psheet[i].column_dimensions['D'].width = 20
            psheet[i].column_dimensions['E'].width = 50
            psheet[i].column_dimensions['F'].width = 20
            psheet[i].column_dimensions['G'].width = 20
            psheet[i].column_dimensions['H'].width = 20
            psheet[i].column_dimensions['I'].width = 15
            psheet[i].column_dimensions['J'].width = 15
            psheet[i].column_dimensions['K'].width = 15
            psheet[i].column_dimensions['L'].width = 15
            psheet[i].column_dimensions['M'].width = 30
            psheet[i].column_dimensions['N'].width = 40
            psheet[i].column_dimensions['O'].width = 30
            psheet[i].row_dimensions[1].height = 30
            for rows in psheet[i].iter_rows(min_row=1, max_row=1, min_col=1):
                for cell in rows:
                  cell.fill = PatternFill(fgColor="FFE78200", fill_type = "solid")
    
            
        # Iterate through the two columns and compare data
        
        # Establish Counting Variables
        
        countMatch = countConflict = countOSP = countUnBoth = countUnPO = countUnPPAR = 0
        
        # Establish an empty list to remember conflicted rows
        
        conflict_rows = []
        
        col_1 = column_index_from_string(c7)  # PO / Item Rev
        col_2 = column_index_from_string(c8)  # PPAR Dwg Rev
        col_3 = column_index_from_string(c2)  # Project Num
        
        for i in range(2, len_max_range):   # Full Spreadsheet
        #for i in range(5073, 5118):        # Testing Purposes Only
            
            QCoreApplication.processEvents()
            
            rev_1 = sheet.cell(row=i, column=col_1).value
            rev_2 = sheet.cell(row=i, column=col_2).value
            projN = sheet.cell(row=i, column=col_3).value
            
            # Sort discrepancies for unspecified revisions
            
            if rev_1 == None:
                rev_1 = '-'
            if rev_2 == None:
                    rev_2 = '-'
                    
            # Disengages filter if none applied
            
            if self.fPN == None:
                projN = self.fPN
            else:
                self.fPN = self.fPN.strip()
            
            # General Analysis
            
            if self.fPN == projN:
                if not any([None in [rev_1, rev_2], '-' in [rev_1, rev_2], '-OSP' in rev_1]):
                    if rev_1 == rev_2:
            #            print '#', i, 'MATCH', rev_1, rev_2
                        countMatch = countMatch + 1
                        for z in range(0,len(c)):
                            psheet[2].cell(row=countMatch+1,column=z+1).value = sheet.cell(row=i,column=column_index_from_string(c[z])).value
                            psheet[2].cell(row=countMatch+1, column=z+1).font = Font(name = 'Bahnschrift', size="12")
                            psheet[2].cell(row=countMatch+1, column = z+1).alignment = Alignment(horizontal='left', vertical='center')
                            for n in range(1,4):
                                psheet[2].row_dimensions[countMatch+1].height = 20
                        
                    else:
                        countConflict = countConflict + 1
                        for z in range(0,len(c)):
                            psheet[1].cell(row=countConflict+1,column=z+1).value = sheet.cell(row=i,column=column_index_from_string(c[z])).value
                            psheet[1].cell(row=countConflict+1, column=z+1).font = Font(name = 'Bahnschrift', size="12")
                            psheet[1].cell(row=countConflict+1, column = z+1).alignment = Alignment(horizontal='left', vertical='center')
                            for n in range(1,4):
                                psheet[1].row_dimensions[countConflict+1].height = 20
    #                    psheet[1].append((cell.value for cell in sheet[i][min_col-1:max_col]))
                        conflict_rows.append(i)
                        poNo = sheet.cell(row=i, column=c4).value
                        pparID = sheet.cell(row=i, column=c9).value
                        print 'CONFLICT ' + str(countConflict) +' <ROW ' + str(i) + '>\n'
                        print '   PO Number: '+ str(poNo)
                        print '   PPAR ID: ' + str(pparID)
                        print '   Rev Discrepancy: PO Rev '+ rev_1 + ', PPAR Rev ' + rev_2 + '\n\n'
                        
                # For OSP PO cases
                        
                elif rev_1 == '-OSP':
            #        print '#', i, 'OSP PO', rev_1, rev_2
                    countOSP = countOSP + 1
                    for z in range(0,len(c)):
                        psheet[3].cell(row=countOSP+1,column=z+1).value = sheet.cell(row=i,column=column_index_from_string(c[z])).value
                        psheet[3].cell(row=countOSP+1, column=z+1).font = Font(name = 'Bahnschrift', size="12")
                        psheet[3].cell(row=countOSP+1, column = z+1).alignment = Alignment(horizontal='left', vertical='center')
                        for n in range(1,4):
                            psheet[3].row_dimensions[countOSP+1].height = 20
    #                psheet[3].append((cell.value for cell in sheet[i][min_col-1:max_col]))
                    
                
                # For cases where revision letter/number is unspecified
                
                else:
                    if rev_1 == rev_2:
            #            print '#', i, 'Unspecified PPAR and PO', rev_1, rev_2
                        countUnBoth = countUnBoth + 1
                    elif rev_1 == '-':
            #            print '#', i, 'Unspecified PO', rev_1, rev_2
                        countUnPO = countUnPO + 1
                    else:
            #            print '#', i, 'Unspecified PPAR', rev_1, rev_2
                        countUnPPAR = countUnPPAR + 1
                
                
                progress = int((float(i) / len_max_range)*100)
    #            print i, progress
                self.percent.emit(str(progress+1)+'%')
        
        
        # Generate Final Report of Analysis
        
        print '\nFINAL OPEN EBS PO ANALYSIS\n'
        print '   TOTAL MATCHES:         ' + str(countMatch)            
        print '   TOTAL CONFLICTS:       ' + str(countConflict)
        print '   TOTAL OSPs:            ' + str(countOSP)
        print '   TOTAL UNSPECIFIED:     ' + str(countUnBoth + countUnPO + countUnPPAR)
        print '      BOTH:               ' + str(countUnBoth)
        print '      PO:                 ' + str(countUnPO)
        print '      PPAR:               ' + str(countUnPPAR)
        print '\n\n   CONFLICTING ROWS:\n', str(conflict_rows)
        
        
        # Format the Overview Section
        
        mr = [psheet[1].max_row,psheet[2].max_row,psheet[3].max_row]
        
        for i in range(1,4):
            for rows in psheet[i].iter_rows(min_row=2, max_row=mr[i-1], min_col=7, max_col=8):
                for cell in rows:
                  cell.fill = PatternFill(fgColor="fff2dabb", fill_type = "solid")

        
        # Merge Cells
        psheet[0].merge_cells(start_row=2, end_row=2, start_column=2, end_column=3)
        psheet[0].merge_cells(start_row=10, end_row=10, start_column=2, end_column=3)
        # Specify Column Heights
        psheet[0].column_dimensions['B'].width = 70
        psheet[0].column_dimensions['C'].width = 100
        # Specify Row Heights
        psheet[0].row_dimensions[1].height = 30
        for i in range(2,10):
            psheet[0].row_dimensions[i].height = 40
        psheet[0].row_dimensions[10].height = 60
        self.set_border(psheet[0],"B2:C9")
        # Font
        psheet[0].cell(row=2, column = 2).font = Font(name = 'Bahnschrift', size="20")
        psheet[0]['B2'].alignment = Alignment(horizontal='center', vertical='center')
        for i in range(3,10):
            psheet[0].cell(row=i, column = 2).font = Font(name = 'Bahnschrift', size="16")
            psheet[0].cell(row=i, column = 2).alignment = Alignment(horizontal='left', vertical='center')
            psheet[0].cell(row=i, column = 3).font = Font(name = 'Bahnschrift', size="16")
            psheet[0].cell(row=i, column = 3).alignment = Alignment(horizontal='center', vertical='center')
        psheet[0].cell(row=10, column = 2).font = Font(name = 'Bahnschrift', size="20", color="FFE78200")
        psheet[0]['B10'].alignment = Alignment(horizontal='center', vertical='center')
        # Cell Background Fills
        psheet[0]['B2'].fill = PatternFill(fgColor="FFE78200", fill_type = "solid")
        psheet[0]['B4'].fill = PatternFill(fgColor="fff2dabb", fill_type = "solid")
        psheet[0]['C4'].fill = PatternFill(fgColor="fff2dabb", fill_type = "solid")
        psheet[0]['B7'].fill = PatternFill(fgColor="fff5f5f5", fill_type = "solid")
        psheet[0]['B8'].fill = PatternFill(fgColor="fff5f5f5", fill_type = "solid")
        psheet[0]['B9'].fill = PatternFill(fgColor="fff5f5f5", fill_type = "solid")
        psheet[0]['C7'].fill = PatternFill(fgColor="fff5f5f5", fill_type = "solid")
        psheet[0]['C8'].fill = PatternFill(fgColor="fff5f5f5", fill_type = "solid")
        psheet[0]['C9'].fill = PatternFill(fgColor="fff5f5f5", fill_type = "solid")
        # Specify Cell Values
        now = datetime.now()
        psheet[0]['B2'] = 'Final Open EBS PO Analysis ' + now.strftime("%m/%d/%Y")
        psheet[0]['B3'] = '    Total Matches'
        psheet[0]['C3'] = countMatch
        psheet[0]['B4'] = '    Total Conflicts'
        psheet[0]['C4'] = countConflict
        psheet[0]['B5'] = '    Total OSPs'
        psheet[0]['C5'] = countOSP
        psheet[0]['B6'] = '    Total Unspecified'
        psheet[0]['C6'] = countUnBoth + countUnPO + countUnPPAR
        psheet[0]['B7'] = '              Both'
        psheet[0]['C7'] = countUnBoth
        psheet[0]['B8'] = '              PO'
        psheet[0]['C8'] = countUnPO
        psheet[0]['B9'] = '              PPAR'
        psheet[0]['C9'] = countUnPPAR
        psheet[0]['B10'] = 'Autogenerated by PXL on ' + now.strftime("%m/%d/%Y, %H:%M:%S")
        
        
        
        pwb.save(newUrl)
        print 'Saved to: ' + newUrl
        self.percent.emit('Opening')
        os.startfile(newUrl)
            
        
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        side = Side(border_style='thin', color="FF000000")
    
        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side
    
                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border
        

# Window UI

class App(QWidget):
    
    font = QFont()
    font.setFamily("Bahnschrift")
    head = "<font face='Bahnschrift' size=8 color=gray align='center'>"
    tail = "</font>"
    head2 = "<font face='Bahnschrift' size=4 color=gray align='center'>"
    
    filterPN = pyqtSignal(str)

    def __init__(self):
        super(App,self).__init__()
        self.title = 'PXL'
#        self.setStyleSheet('background:#f5f5f5')
        self.left = 10
        self.top = 10
        self.width = 840
        self.height = 680
        self.initUI()
        self.mode = pyqtSignal(int)
            
    def initUI(self):
        self.setWindowTitle(self.title)
        self.setWindowIcon(QIcon(icon))
        self.setGeometry(self.left, self.top, self.width, self.height)
        
        # Banner
        
        self.logo = QLabel(self)
        pixmap = QPixmap(logo)
        self.logo.setPixmap(pixmap)
        self.cb = QComboBox(self)
        self.cb.setMaximumWidth(200)
        self.cb.addItem('PO/PPAR Revision Check')
        self.cb.setStyleSheet('''   QComboBox{font-family:Bahnschrift SemiLight; background-color:transparent; color:gray; border-color:darkgray; selection-background-color:transparent; padding: 0px 0px 0px 0px; padding-right:10px;}
                                    QComboBox:down-arrow{color:rgb(231,130,0);}
                                    QAbstractItemView {color:gray; background-color:lightgray; selection-background-color: rgb(231,130,0); width:100px; text-align:right;}
                                    QAbstractItemView:item{height:20px}
                                    QComboBox:hover{background-color:transparent; color: rgb(231,130,0)}''')
        
        self.cb.currentIndexChanged.connect(self.mode)
        
        self.fbtn = QPushButton('Filter')
        self.fbtn.setStyleSheet(''' QPushButton{font-family:Bahnschrift SemiLight; background-color:transparent; color:gray; border-color:darkgray; selection-background-color:transparent; padding: 0px 0px 0px 0px; padding-right:10px;}
                                    QPushButton:hover{background-color:transparent; color: rgb(231,130,0)}''')
        self.fbtn.clicked.connect(self.getFilter)
        
        self.settingsLayout = QHBoxLayout()
        self.settingsLayout.addWidget(self.fbtn)        
        self.settingsLayout.addWidget(self.cb)
        self.settingsLayout.setAlignment(Qt.AlignRight)
        self.settingsLayout.setSpacing(40)
        
        
        self.bannerLayout = QHBoxLayout()
        self.bannerLayout.addWidget(self.logo)
        self.bannerLayout.addLayout(self.settingsLayout)
        
        
        # Framing Setup
        
        self.verticalLayout_2 = QVBoxLayout(self)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        
        self.verticalLayout = QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        
        
        self.horizontalLayout_2 = QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        
        self.horizontalLayout = QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.horizontalLayout_2.addLayout(self.horizontalLayout)
        
        # Consolidate Layouts
        
        self.verticalLayout.addLayout(self.bannerLayout)
        
        self.frame_2 = QFrame()
        self.frame_2.setFrameShape(QFrame.StyledPanel)
        self.frame_2.setFrameShadow(QFrame.Raised)
        self.frame_2.setObjectName("frame_2")
        self.frame_2.setStyleSheet('border:transparent; color:#828282; font-family:Bahnschrift;')
        
        self.horizontalLayout_3 = QHBoxLayout(self.frame_2)
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        
        self.verticalLayout_5 = QVBoxLayout()
        self.verticalLayout_5.setObjectName("verticalLayout_5")
        
        # Drag and Drop Section
        
        self.labelTarget = Box()
        self.labelTarget.working.connect(self.busy)
        self.labelTarget.setMinimumHeight(150)
        self.labelTarget.setAcceptDrops(True)   
        self.verticalLayout_5.addWidget(self.labelTarget)
        self.horizontalLayout_3.addLayout(self.verticalLayout_5)
        self.verticalLayout.addWidget(self.frame_2)
        
        # Percent Label
        
#        self.label = QLabel()
#        self.label.setText("69%")
#        self.label.setStyleSheet("color:gray;" "font: bold 14pt 'Bahnschrift'")
##        width, height = self.frameGeometry().width(), self.frameGeometry().height()
##        print width, height
##        self.label.move(width/2-27,height/2-60)
#        self.label.show()
        
        
        # Bottom Section
        
        self.frame_3 = QFrame()
        self.frame_3.setFrameShape(QFrame.StyledPanel)
        self.frame_3.setFrameShadow(QFrame.Raised)
        self.frame_3.setObjectName("frame_3")
        self.frame_3.setStyleSheet('color:#828282;')
        self.frame_3.setMaximumHeight(100)
        self.verticalLayout_4 = QVBoxLayout(self.frame_3)
        
        self.verticalLayout_4.setObjectName("verticalLayout_4")
        
        self.verticalLayout_3 = QVBoxLayout()
        
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        
        self.labelQuestion = QLabel(self.frame_3)
        self.labelQuestion.setObjectName("labelQuestion")
        self.labelQuestion.setText("   Save PXL analysis to:")
        self.labelQuestion.setStyleSheet('font-family:bahnschrift;')
        self.labelQuestion.setFont(self.font)
        
        self.verticalLayout_3.addWidget(self.labelQuestion)
        
        self.frame_4 = QFrame(self.frame_3)
        self.frame_4.setFrameShape(QFrame.StyledPanel)
        self.frame_4.setFrameShadow(QFrame.Raised)
        self.frame_4.setObjectName("frame_4")
        
        
        self.horizontalLayout_5 = QHBoxLayout(self.frame_4)
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")
        
        savePath = readTxt()
        self.inputDirectory = QLineEdit(self.frame_4)
        self.inputDirectory.setObjectName("inputDirectory")
        self.inputDirectory.setText(savePath)
        self.inputDirectory.setStyleSheet("QLineEdit { padding-bottom: 2px; padding-left: 2px; font-family:bahnschrift;}")
        
        self.horizontalLayout_5.addWidget(self.inputDirectory)
        
        self.buttonBrowse = QPushButton(self.frame_4)
        self.buttonBrowse.setObjectName("buttonBrowse")
        self.buttonBrowse.setText("Browse")
        self.buttonBrowse.setStyleSheet('font-family:bahnschrift;')
        self.buttonBrowse.clicked.connect(self.openFolder)
        
        self.horizontalLayout_5.addWidget(self.buttonBrowse)
        self.horizontalLayout_4 = QHBoxLayout()
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        self.horizontalLayout_5.addLayout(self.horizontalLayout_4)
        
        self.verticalLayout_3.addWidget(self.frame_4)
        self.verticalLayout_4.addLayout(self.verticalLayout_3)
        self.verticalLayout.addWidget(self.frame_3)
        self.verticalLayout_2.addLayout(self.verticalLayout)
        
        
        # Window
        
        qtRectangle = self.frameGeometry()
        centerPoint = QDesktopWidget().availableGeometry().center()
        qtRectangle.moveCenter(centerPoint)
        self.move(qtRectangle.topLeft())
        self.show()
    
    def openFolder(self):
         options = QFileDialog.Options()
         options |= QFileDialog.DirectoryOnly
         folderName = QFileDialog.getExistingDirectory(self, 'Select Folder')
         if folderName:
             savePath = folderName
             self.inputDirectory.setText(savePath)
             f = open(txt, "w")
             f.write(savePath)
             f.close()
             
    def mode(self):
        modeIndex = self.cb.currentIndex()
        print modeIndex
        
    def busy(self, status):
        if status == 1:
            self.buttonBrowse.setEnabled(False)
            self.inputDirectory.setEnabled(False)
            self.fbtn.setEnabled(False)
            self.cb.setEnabled(False)
        elif status == 0:
            self.buttonBrowse.setEnabled(True)
            self.inputDirectory.setEnabled(True)
            self.fbtn.setEnabled(True)
            self.cb.setEnabled(True)
            
        
    def getFilter(self):
        Box.fPN, okPressed = QInputDialog.getText(self, "Apply Filter",self.head2+"<br>Specify the applicable project number          <br>"+self.tail, QLineEdit.Normal, Box.fPN)
        if okPressed and Box.fPN != '':
            self.labelTarget.unfade()
            self.labelTarget.setText(self.head+"<b>Filter Updated:</b> analysis will be for the project number <b>'"+Box.fPN+"'</b>"+self.tail)
        else:
            self.labelTarget.unfade()
            self.labelTarget.setText(self.head+"<b>Filter Updated:</b> there is no filter currently in use."+self.tail)
            Box.fPN = None
        print 'Filter is set to:',Box.fPN
        Box.resetTimer = QTimer()
        Box.resetTimer.setInterval(4000)
        Box.resetTimer.setSingleShot(True)
        Box.resetTimer.timeout.connect(self.labelTarget.reset)
        Box.resetTimer.start()

             
    def fade(self, widget):
        self.effect = QGraphicsOpacityEffect()
        widget.setGraphicsEffect(self.effect)
    
        self.animation = QPropertyAnimation(self.effect, b"opacity")
        self.animation.setDuration(1000)
        self.animation.setStartValue(1)
        self.animation.setEndValue(0)
        self.animation.start()

    def unfade(self, widget):
        self.effect = QGraphicsOpacityEffect()
        widget.setGraphicsEffect(self.effect)
    
        self.animation = QPropertyAnimation(self.effect, b"opacity")
        self.animation.setDuration(1000)
        self.animation.setStartValue(0)
        self.animation.setEndValue(1)
        self.animation.start()
             

    
if __name__ == '__main__':
    app = QApplication(sys.argv)
    db = QFontDatabase.addApplicationFont(fnt)
#    app.setStyleSheet('font-family:bahnschrift;')
    ex = App()
    sys.exit(app.exec_())